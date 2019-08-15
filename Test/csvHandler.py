import os
import math
import csv
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
import shutil

def csv2xl(csvFileName, xlFileName, xlSheetName, kommunekode, macroEnebeled):
    os.listdir('.')
    workingFileName = xlFileName+kommunekode+'.xlsx'
    print("Writing "+csvFileName+" to "+workingFileName)
    shutil.copyfile(xlFileName+'.xlsx', workingFileName)
    data = pd.read_csv(csvFileName, encoding='latin1', header=0, quotechar='"', delimiter=";")

    if macroEnebeled:
        book = load_workbook(filename=workingFileName, read_only=False, keep_vba=True)
    else:
        book = load_workbook(workingFileName)

    writer = pd.ExcelWriter(workingFileName, engine='openpyxl')
    writer.book = book

    data.to_excel(writer, sheet_name=xlSheetName)
    writer.save()
    writer.close()
    print("Saved " + workingFileName + " successfully")

    filename_macro = workingFileName + '.xlsm'
    workbook = writer.book
    workbook.filename = filename_macro
    workbook.add_vba_project('vbaProject.bin')
    writer.save()

def xl2csv(csvFileName, xlsxFileName, xlsxSheetName):
    print("Writing "+xlsxFileName+" to "+csvFileName)
    os.listdir('.')
    data_xls = pd.read_excel(xlsxFileName, xlsxSheetName, index_col=1, encoding='utf-8')
    data_xls.to_csv(csvFileName, encoding='utf-8-sig', header=True, index=False, quotechar='"', sep=";")



